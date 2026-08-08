"""
ハイキュー!! バボカ!! BREAK — 全カードデータ + 画像ダウンロード & Excel 生成スクリプト
================================================================
動作環境: Python 3.8+
必要ライブラリ: pip install requests pillow openpyxl
================================================================
"""

import requests
import shutil
import time
import json
import os
import sys
from pathlib import Path
from io import BytesIO
from datetime import datetime

# ─── 設定 ───────────────────────────────────────────────────────
BASE_API   = "https://www.takaratomy.co.jp/products/haikyuvobacabreak/cardlist/itemsearch.php"
IMG_BASE   = "https://www.takaratomy.co.jp/products/haikyuvobacabreak/cardlist/card"
REFERER    = "https://www.takaratomy.co.jp/products/haikyuvobacabreak/cardlist/"
HEADERS    = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Referer":    REFERER,
}
OUTPUT_DIR   = Path("haikyuu_output")
IMG_DIR      = OUTPUT_DIR / "images"
SITE_IMG_DIR = Path("images")  # 網站實際讀圖的位置（index.html/game.html 都是相對路徑 images/）——
                                # haikyuu_output/images/ 只是下載暫存區，不會被網站讀到
JSON_PATH    = OUTPUT_DIR / "all_cards.json"
EXCEL_PATH   = OUTPUT_DIR / "haikyuu_cards_with_images.xlsx"
REQUEST_WAIT = 0.4   # API リクエスト間隔(秒)
IMG_WAIT     = 0.3   # 画像リクエスト間隔(秒)
IMG_THUMB_H  = 80    # Excel 内サムネ高さ(px)
IMG_THUMB_W  = 58    # Excel 内サムネ幅(px)
ROW_HEIGHT   = 62    # Excel 行高(pt)
CARD_DETAIL_URL = "https://www.takaratomy.co.jp/products/haikyuvobacabreak/cardlist/card.html"
QA_WAIT      = 0.5   # Q&A リクエスト間隔(秒)
QA_PATH      = OUTPUT_DIR / "qa_data.json"

# ─── Step 1: API 全ページ取得 ─────────────────────────────────────
def fetch_all_cards() -> list[dict]:
    print("=" * 60)
    print("【Step 1】カードデータ取得開始")
    print("=" * 60)

    all_items: list[dict] = []
    page = 1
    total_declared = None

    while True:
        try:
            res = requests.get(BASE_API, params={"p": page}, headers=HEADERS, timeout=15)
            res.raise_for_status()
            data = res.json()
        except Exception as e:
            print(f"  [ERROR] ページ {page} 取得失敗: {e}")
            break

        items = data.get("items", [])
        if total_declared is None:
            total_declared = data.get("count", 0)
            print(f"  API宣言総数: {total_declared} バリアント")

        if not items:
            print(f"  第 {page} ページ: 空データ → 終了")
            break

        all_items.extend(items)
        print(f"  第 {page:>3} ページ: {len(items):>3} 件取得 | 累計 {len(all_items):>4} / {total_declared}")

        # 終了判定: 本ページが空 OR 累計が宣言総数に達した
        if total_declared > 0 and len(all_items) >= total_declared:
            print(f"  宣言総数 {total_declared} に到達 → 終了")
            break

        page += 1
        time.sleep(REQUEST_WAIT)

    # ID で重複除去（同一IDは後勝ち）
    unique = {item["ID"]: item for item in all_items}
    print(f"\n  取得完了: {len(all_items)} バリアント / 重複除去後: {len(unique)} ユニーク")
    return list(unique.values())


# ─── Step 2: 画像 URL 構築 ───────────────────────────────────────
def build_image_url(card: dict) -> str:
    """
    URL パターン: {IMG_BASE}/{card_no}-{image_end}.webp
    確認済みパターン:
      HV-P01-014-R.webp   (Rare)
      HV-P01-014-RP.webp  (Rare Parallel)
      HV-P01-018-I.webp   (頂)
      HV-P01-018-IP.webp  (頂 Parallel)
      HV-P01-017-H.webp   (秘 Secret)
      HV-P01-022-S.webp   (Super)
      HV-P01-045-SP.webp  (Super Parallel)
      HV-P01-015-N.webp   (Normal)
      HV-P01-015-NP.webp  (Normal Parallel)
      HV-D01-001-D.webp   (Starter Deck)
      HV-PR-001-P.webp    (Promo)
    """
    card_no   = card.get("card_no", "")
    image_end = card.get("image_end", "")
    return f"{IMG_BASE}/{card_no}-{image_end}.webp"


# ─── Step 3: 画像ダウンロード ──────────────────────────────────────
def download_images(cards: list[dict]) -> dict[str, Path]:
    """画像をダウンロードして {ID: local_path} を返す"""
    print("\n" + "=" * 60)
    print("【Step 2】画像ダウンロード開始")
    print("=" * 60)

    IMG_DIR.mkdir(parents=True, exist_ok=True)
    id_to_path: dict[str, Path] = {}
    success = fail = skip = 0

    for i, card in enumerate(cards, 1):
        card_id  = card.get("ID", f"unknown_{i}")
        card_no  = card.get("card_no", "")
        img_end  = card.get("image_end", "")
        filename = f"{card_no}-{img_end}.webp"
        local    = IMG_DIR / filename

        # 既存ファイルはスキップ
        if local.exists() and local.stat().st_size > 1000:
            id_to_path[card_id] = local
            skip += 1
            if i % 50 == 0:
                print(f"  [{i:>4}/{len(cards)}] スキップ済 {skip} 件")
            continue

        url = build_image_url(card)
        try:
            r = requests.get(url, headers=HEADERS, timeout=10)
            if r.status_code == 200:
                local.write_bytes(r.content)
                id_to_path[card_id] = local
                success += 1
                print(f"  [{i:>4}/{len(cards)}] ✅ {filename} ({len(r.content)//1024}KB)")
            else:
                fail += 1
                print(f"  [{i:>4}/{len(cards)}] ❌ {filename} → HTTP {r.status_code}")
        except Exception as e:
            fail += 1
            print(f"  [{i:>4}/{len(cards)}] ❌ {filename} → {e}")

        time.sleep(IMG_WAIT)

    print(f"\n  ✅ 成功: {success}  ⏭ スキップ: {skip}  ❌ 失敗: {fail}")
    return id_to_path


# ─── Step 2.5: 網站用画像ディレクトリへ同期 ──────────────────────────
def sync_images_to_site(cards: list[dict]) -> int:
    """haikyuu_output/images/ からサイトが実際に読む images/ へ新規カード画像をコピーする。
    index.html/game.html 等は相対パス images/ を直接参照しており、haikyuu_output/images/ は
    ダウンロード用の一時置き場に過ぎない——このコピーを忘れると新カードの画像がサイト上で
    表示されない（2026-08 に発見・修正）。"""
    print("\n" + "=" * 60)
    print("【Step 2.5】サイト用 images/ への同期")
    print("=" * 60)

    SITE_IMG_DIR.mkdir(parents=True, exist_ok=True)
    copied = 0
    for card in cards:
        filename = f"{card.get('card_no', '')}-{card.get('image_end', '')}.webp"
        src = IMG_DIR / filename
        dst = SITE_IMG_DIR / filename
        if not src.exists():
            continue
        if dst.exists() and dst.stat().st_size == src.stat().st_size:
            continue
        shutil.copyfile(src, dst)
        copied += 1
        print(f"  📋 {filename}")

    print(f"\n  同期完了: 新規/更新 {copied} 件")
    return copied


# ─── Step 1.5: Q&A データ取得 ─────────────────────────────────────
import re as _re


def _strip_tags(text: str) -> str:
    """HTML タグを除去してテキストを返す"""
    return _re.sub(r"<[^>]+>", "", text).strip()


QA_API = "https://www.takaratomy.co.jp/products/haikyuvobacabreak/rules/itemsearch.php"


def fetch_qa_data(cards: list[dict], only_missing: list[str] | None = None) -> dict:
    """only_missing を渡すと、そのcard_noだけ取得する（増分取得。fetch_new_qa.py 用）。
    渡さない場合は cards 全体から重複除去した card_no 全部を取得する（従来通り）。"""
    print("\n" + "=" * 60)
    print("【Step 1.5】Q&A データ取得開始")
    print("=" * 60)

    if only_missing is not None:
        card_nos = sorted(set(only_missing))
    else:
        # card_no ごとに重複除去
        card_nos = sorted(set(c.get("card_no", "") for c in cards if c.get("card_no")))
    total = len(card_nos)
    print(f"  対象 card_no: {total} 枚")

    qa_data: dict[str, list] = {}

    for idx, card_no in enumerate(card_nos, 1):
        try:
            res = requests.get(
                QA_API,
                params={"cn": card_no},
                headers=HEADERS,
                timeout=15,
            )
            res.raise_for_status()
            data = res.json()
        except Exception as e:
            print(f"  [ERROR] {card_no} 取得失敗: {e}")
            qa_data[card_no] = []
            time.sleep(QA_WAIT)
            continue

        if int(data.get("count", 0)) == 0:
            qa_data[card_no] = []
        else:
            items = []
            for group in data.get("items", {}).values():
                for qa in group.get("items", []):
                    def clean(t: str) -> str:
                        return (t or "").strip().replace("\r\n", "\n").replace("\r", "\n")
                    items.append({
                        "id":       qa.get("id", ""),
                        "date":     (qa.get("sdate", "") or "")[:10],
                        "question": clean(qa.get("que", "")),
                        "answer":   clean(qa.get("ans", "")),
                    })
            qa_data[card_no] = items

        if idx % 10 == 0 or idx == total:
            print(f"  進捗: {idx}/{total}  ({card_no}  Q&A: {len(qa_data[card_no])} 件)")

        time.sleep(QA_WAIT)

    total_qa = sum(len(v) for v in qa_data.values())
    cards_with_qa = sum(1 for v in qa_data.values() if v)
    print(f"\n  取得完了: {cards_with_qa}/{total} 枚に Q&A あり  合計 {total_qa} 件")
    return qa_data


# ─── Step 4: Excel 生成 ───────────────────────────────────────────
def build_excel(cards: list[dict], id_to_path: dict[str, Path]):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.formatting.rule import ColorScaleRule
        from PIL import Image as PILImage
    except ImportError as e:
        print(f"\n[ERROR] ライブラリ不足: {e}")
        print("  pip install openpyxl pillow  を実行してください")
        return

    print("\n" + "=" * 60)
    print("【Step 3】Excel 生成開始")
    print("=" * 60)

    wb = Workbook()

    # ─── レアリティ表示名マッピング ───
    RARITY_LABEL = {
        "H": "秘", "I": "頂", "IP": "頂P",
        "S": "S", "SP": "SP", "R": "R", "RP": "RP",
        "N": "N", "NP": "NP", "D": "D", "P": "P",
    }
    RARITY_COLOR = {
        "H":  "FFD700",  # 金
        "I":  "C084FC",  # 紫
        "IP": "A855F7",
        "S":  "60A5FA",  # 青
        "SP": "3B82F6",
        "R":  "34D399",  # 緑
        "RP": "10B981",
        "N":  "D1D5DB",  # 灰
        "NP": "9CA3AF",
        "D":  "FCD34D",  # 黄
        "P":  "FB923C",  # オレンジ
    }
    SCHOOL_COLOR = {
        "烏野": "E96F24", "音駒": "D60527", "稲荷崎": "6B21A8",
        "白鳥沢": "1D4ED8", "伊達工業": "065F46", "青葉城西": "0369A1",
        "梟谷": "B45309", "鴎台": "7C3AED",
    }

    def get_school(affiliation: str) -> str:
        for s in SCHOOL_COLOR:
            if s in affiliation:
                return s
        return "その他"

    # ─── Sheet 1: 全カード一覧（画像付き）─────────────────────────
    ws = wb.active
    ws.title = "全カード一覧"

    HDR_FILL  = PatternFill("solid", fgColor="1E293B")
    HDR_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    DATA_FONT = Font(name="Calibri", size=9)
    TITLE_FONT= Font(name="Calibri", bold=True, size=14, color="1E293B")
    THIN      = Side(style="thin", color="D1D5DB")
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT      = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)

    # タイトル行
    ws.column_dimensions["A"].width = 3
    ws.row_dimensions[1].height = 8
    ws.merge_cells("B2:P2")
    title_cell = ws["B2"]
    title_cell.value = f"ハイキュー!! バボカ!! BREAK — 全カードリスト  ({len(cards)} バリアント)"
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER
    ws.row_dimensions[2].height = 28

    ws.merge_cells("B3:P3")
    sub_cell = ws["B3"]
    sub_cell.value = f"生成日時: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Source: takaratomy.co.jp/products/haikyuvobacabreak/cardlist/"
    sub_cell.font = Font(name="Calibri", size=8, color="6B7280")
    sub_cell.alignment = LEFT
    ws.row_dimensions[3].height = 16

    # ヘッダー定義
    COLS = [
        ("画像",     10),
        ("カードNo.", 16),
        ("キャラ名",  14),
        ("よみがな",  16),
        ("所属",     14),
        ("学校",     10),
        ("レアリティ", 10),
        ("ポジション", 12),
        ("サーブ",    7),
        ("ブロック",   7),
        ("レシーブ",   7),
        ("トス",      7),
        ("アタック",   7),
        ("合計",      7),
        ("スキル",    40),
    ]

    HEADER_ROW = 4
    ws.row_dimensions[HEADER_ROW].height = 22
    for ci, (col_name, col_w) in enumerate(COLS, 2):
        cell = ws.cell(row=HEADER_ROW, column=ci, value=col_name)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = col_w

    # データ行
    data_start_row = HEADER_ROW + 1
    for ri, card in enumerate(cards, data_start_row):
        ws.row_dimensions[ri].height = ROW_HEIGHT
        card_id   = card.get("ID", "")
        card_no   = card.get("card_no", "")
        img_end   = card.get("image_end", "")
        name      = card.get("name", "")
        name_ruby = card.get("name_ruby", "")
        affil     = card.get("affiliation", "")
        school    = get_school(affil)
        rarity_raw= img_end
        rarity    = RARITY_LABEL.get(img_end, img_end)
        position  = card.get("position_ruby") or card.get("position", "")
        serve     = card.get("serve", "-")
        block     = card.get("block", "-")
        receive   = card.get("receive", "-")
        toss      = card.get("toss", "-")
        attack    = card.get("attack", "-")
        skill     = (card.get("skill") or "").replace("\r\n", " ").strip()
        if skill == "-": skill = ""

        # 数値計算
        def to_num(v):
            try: return int(v)
            except: return 0
        total = to_num(serve) + to_num(block) + to_num(receive) + to_num(toss) + to_num(attack)

        row_data = [
            "",        # 画像プレースホルダー
            card_no, name, name_ruby, affil, school,
            rarity, position,
            serve, block, receive, toss, attack, total,
            skill,
        ]
        for ci, val in enumerate(row_data, 2):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = DATA_FONT
            cell.border = BORDER
            if ci in (2, 7, 8, 9, 10, 11, 12, 13, 14):
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT

            # レアリティ列を色付け
            if ci == 8:
                rc = RARITY_COLOR.get(rarity_raw, "FFFFFF")
                cell.fill = PatternFill("solid", fgColor=rc)
                cell.font = Font(name="Calibri", size=9, bold=True,
                                 color="000000" if rc in ("D1D5DB","9CA3AF","FCD34D","FB923C","FFD700") else "FFFFFF")

            # 合計列を色付け
            if ci == 15 and isinstance(val, int):
                if val >= 15:
                    cell.fill = PatternFill("solid", fgColor="BBF7D0")
                elif val >= 10:
                    cell.fill = PatternFill("solid", fgColor="FEF9C3")

        # 画像埋め込み
        img_path = id_to_path.get(card_id)
        if img_path and img_path.exists():
            try:
                # WebP → PNG へ変換（openpyxl は webp 非対応）
                pil_img = PILImage.open(img_path)
                pil_img.thumbnail((IMG_THUMB_W, IMG_THUMB_H), PILImage.LANCZOS)
                buf = BytesIO()
                pil_img.save(buf, format="PNG")
                buf.seek(0)

                xl_img = XLImage(buf)
                xl_img.width  = IMG_THUMB_W
                xl_img.height = IMG_THUMB_H

                col_b_letter = get_column_letter(2)
                cell_addr = f"{col_b_letter}{ri}"
                ws.add_image(xl_img, cell_addr)
            except Exception as e:
                ws.cell(row=ri, column=2, value="[画像エラー]")
        else:
            ws.cell(row=ri, column=2, value="[なし]")

        if ri % 50 == 0:
            print(f"  Excel 書き込み: {ri - data_start_row + 1}/{len(cards)} 行")

    # Excelテーブル（画像列を除く B4:P末尾は手動フィルター）
    last_row = data_start_row + len(cards) - 1
    ws.auto_filter.ref = f"B{HEADER_ROW}:{get_column_letter(len(COLS)+1)}{last_row}"
    ws.freeze_panes = f"B{data_start_row}"

    # 合計列 ColorScale
    from openpyxl.formatting.rule import ColorScaleRule
    cs_rule = ColorScaleRule(
        start_type="min", start_color="FFFFFF",
        end_type="max",   end_color="22C55E"
    )
    ws.conditional_formatting.add(f"O{data_start_row}:O{last_row}", cs_rule)

    # ─── Sheet 2: 学校別サマリー ──────────────────────────────────
    ws2 = wb.create_sheet("学校別サマリー")
    ws2["B2"] = "学校別カード枚数サマリー"
    ws2["B2"].font = Font(name="Calibri", bold=True, size=13, color="1E293B")
    ws2.row_dimensions[2].height = 24

    s_headers = ["学校", "バリアント数", "ユニークカードNo.", "秘レア数", "頂レア数", "平均合計P"]
    for ci, h in enumerate(s_headers, 2):
        cell = ws2.cell(row=4, column=ci, value=h)
        cell.font = HDR_FONT; cell.fill = HDR_FILL
        cell.alignment = CENTER; cell.border = BORDER
    ws2.row_dimensions[4].height = 20

    from collections import defaultdict
    school_data = defaultdict(list)
    for card in cards:
        school_data[get_school(card.get("affiliation",""))].append(card)

    sr = 5
    for school, sc_cards in sorted(school_data.items(), key=lambda x: -len(x[1])):
        unique_nos = len(set(c["card_no"] for c in sc_cards))
        secret_cnt = sum(1 for c in sc_cards if c.get("image_end") == "H")
        top_cnt    = sum(1 for c in sc_cards if c.get("image_end") in ("I","IP"))
        totals = []
        for c in sc_cards:
            def n(v):
                try: return int(v)
                except: return 0
            t = n(c.get("serve",0))+n(c.get("block",0))+n(c.get("receive",0))+n(c.get("toss",0))+n(c.get("attack",0))
            if t > 0: totals.append(t)
        avg_total = round(sum(totals)/len(totals), 1) if totals else 0

        row_vals = [school, len(sc_cards), unique_nos, secret_cnt, top_cnt, avg_total]
        for ci, v in enumerate(row_vals, 2):
            cell = ws2.cell(row=sr, column=ci, value=v)
            cell.font = DATA_FONT; cell.alignment = CENTER; cell.border = BORDER
            sc = SCHOOL_COLOR.get(school, "F3F4F6")
            if ci == 2:
                cell.fill = PatternFill("solid", fgColor=sc)
                cell.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        sr += 1

    for ci, w in enumerate([12, 12, 16, 10, 10, 12], 2):
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.column_dimensions["A"].width = 3
    ws2.freeze_panes = "B5"

    # ─── 保存 ──────────────────────────────────────────────────────
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(EXCEL_PATH)
    print(f"\n  ✅ Excel 保存完了: {EXCEL_PATH}")
    print(f"     ファイルサイズ: {EXCEL_PATH.stat().st_size // 1024} KB")


# ─── 官方新卡/新QA偵測（見 docs/translation/CONTEXT.md「新卡／新QA偵測」）───
def should_refetch(official_count: int, local_count: int) -> bool:
    """純決策：官方宣告總數跟本地已有張數不一致時，才需要觸發全量重抓。"""
    return official_count != local_count


def is_fetch_complete(fetched_count: int, official_count: int) -> bool:
    """純決策：check_new_cards.py 寫檔前的完整性檢查。全量重抓若中途部分失敗，
    可能只抓到一部分就被誤當完整快照——抓到的張數不足官方宣告總數就視為不完整。"""
    return fetched_count >= official_count


def is_removal_suspicious(removed_count: int, local_count: int, threshold: float = 0.3) -> bool:
    """純決策：下架比例超過門檻（預設 30%）時視為可疑，比較可能是 API 暫時異常
    而不是真的大量下架，check_new_cards.py 應該中止寫檔而非靜默覆寫。"""
    if local_count <= 0:
        return False
    return (removed_count / local_count) > threshold


def fetch_official_count() -> int:
    """只打 API 第一頁讀宣告總數，不做完整分頁抓取（比 fetch_all_cards() 便宜很多）。"""
    res = requests.get(BASE_API, params={"p": 1}, headers=HEADERS, timeout=15)
    res.raise_for_status()
    return int(res.json().get("count", 0))


def check_for_new_cards() -> dict:
    """比對官方宣告總數跟本地 all_cards.json 張數；不一致才觸發全量重抓，
    並用穩定的卡片 variant 層級 ID 欄位（不是 QA API 那個每次重抓會重新編號的 id，
    見 docs/translation/docs/adr/0004）diff 出新增/下架的卡。
    只回傳結果，不自動覆寫 all_cards.json——要不要寫檔留給呼叫端決定。
    """
    local_cards = []
    if JSON_PATH.exists():
        with open(JSON_PATH, encoding="utf-8") as f:
            local_cards = json.load(f)

    official_count = fetch_official_count()
    local_count = len(local_cards)
    has_new = should_refetch(official_count, local_count)

    result = {
        "official_count": official_count,
        "local_count": local_count,
        "has_new": has_new,
        "added_ids": [],
        "removed_ids": [],
        "fetched_cards": None,
    }
    if not has_new:
        return result

    fetched = fetch_all_cards()
    old_ids = {c["ID"] for c in local_cards if "ID" in c}
    new_ids = {c["ID"] for c in fetched if "ID" in c}
    result["added_ids"] = sorted(new_ids - old_ids)
    result["removed_ids"] = sorted(old_ids - new_ids)
    result["fetched_cards"] = fetched
    return result


def _load_or_fetch(path, fetch_fn, interactive: bool, label: str):
    """既存ファイルがあれば（非interactiveなら問答無用で）流用、なければ fetch_fn() で取得して保存。
    JSON/QA 両ステップが同じ形をしていたので一つにまとめた（interactive パラメータ導入時の重複）。"""
    if path.exists():
        if interactive:
            ans = input(f"\n既存の{label}が見つかりました ({path})。\n再取得しますか？ [y/N]: ").strip().lower()
        else:
            ans = "n"
        if ans != "y":
            print(f"  既存{label}を使用します")
            with open(path, encoding="utf-8") as f:
                return json.load(f)

    data = fetch_fn()
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  {label}を保存: {path}")
    return data


# ─── メイン ───────────────────────────────────────────────────────
def main(interactive: bool = True):
    print(f"\nハイキュー!! バボカ!! BREAK カードダウンロードツール")
    print(f"開始時刻: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Step 1: カードデータ取得
    cards = _load_or_fetch(JSON_PATH, fetch_all_cards, interactive, "JSON")

    # Step 1.5: Q&A データ取得
    qa_data = _load_or_fetch(QA_PATH, lambda: fetch_qa_data(cards), interactive, "QAデータ")

    # Step 2: 画像ダウンロード
    id_to_path = download_images(cards)

    # Step 2.5: サイト用 images/ への同期
    sync_images_to_site(cards)

    # Step 3: Excel 生成
    build_excel(cards, id_to_path)

    total_qa_count = sum(len(v) for v in qa_data.values())
    cards_with_qa_count = sum(1 for v in qa_data.values() if v)
    print(f"\n{'='*60}")
    print(f"完了！ 出力フォルダ: {OUTPUT_DIR.resolve()}")
    print(f"  📄 JSON : {JSON_PATH.name}")
    print(f"  ❓ QA   : {QA_PATH.name} ({cards_with_qa_count} 枚 / 合計 {total_qa_count} 件)")
    print(f"  🖼  画像 : {IMG_DIR.name}/ ({len(list(IMG_DIR.glob('*.webp')))} ファイル)")
    print(f"  📊 Excel: {EXCEL_PATH.name}")
    print(f"終了時刻: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    main()
